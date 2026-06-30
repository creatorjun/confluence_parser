# to_excel.py
from __future__ import annotations

from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config import ROOT_PAGE_ID
from confluence_client import collect_pages
from html_cleaner import clean, is_code_table, extract_code_lines

OUTPUT_FILE = "output.xlsx"

_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
_SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_SUBHEADER_FONT = Font(name="맑은 고딕", bold=True, color="1F3864", size=10)
_BODY_FONT = Font(name="맑은 고딕", size=10)
_CODE_FONT = Font(name="Consolas", size=9, color="333333")
_CODE_FILL = PatternFill("solid", fgColor="F4F4F4")
_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _write_heading(ws, row: int, text: str, depth: int) -> int:
    col_fill = _HEADER_FILL if depth <= 1 else _SUBHEADER_FILL
    col_font = _HEADER_FONT if depth <= 1 else _SUBHEADER_FONT
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = col_fill
    cell.font = col_font
    cell.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    return row + 1


def _write_table(ws, row: int, table_tag: Tag) -> int:
    rows = table_tag.find_all("tr")
    for r_idx, tr in enumerate(rows):
        cells = tr.find_all(["td", "th"])
        col = 1
        for cell_tag in cells:
            text = cell_tag.get_text(strip=True).replace("\u00a0", " ")
            colspan = int(cell_tag.get("colspan", 1))
            is_header = cell_tag.name == "th" or r_idx == 0
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = _HEADER_FONT if is_header else _BODY_FONT
            cell.fill = _SUBHEADER_FILL if is_header else PatternFill()
            cell.border = _BORDER
            cell.alignment = _WRAP
            if colspan > 1:
                ws.merge_cells(
                    start_row=row, start_column=col,
                    end_row=row, end_column=col + colspan - 1
                )
            col += colspan
        row += 1
    return row + 1


def _write_code(ws, row: int, text: str) -> int:
    for line in text.split("\n"):
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = _CODE_FONT
        cell.fill = _CODE_FILL
        cell.alignment = Alignment(wrap_text=False, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    return row + 1


def _write_paragraph(ws, row: int, text: str) -> int:
    if not text.strip():
        return row
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _BODY_FONT
    cell.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    return row + 1


def _process_node(ws, node, row: int) -> int:
    from bs4 import NavigableString

    if isinstance(node, NavigableString):
        return row
    tag = node.name
    if tag is None:
        return row

    if tag in ("h1", "h2", "h3", "h4", "h5", "h6"):
        row = _write_heading(ws, row, node.get_text(strip=True), int(tag[1]) - 1)
    elif tag == "p":
        text = node.get_text().replace("\u00a0", " ").strip()
        row = _write_paragraph(ws, row, text)
    elif tag in ("pre", "code"):
        row = _write_code(ws, row, node.get_text())
    elif tag == "table":
        if is_code_table(node):
            row = _write_code(ws, row, extract_code_lines(node))
        else:
            row = _write_table(ws, row, node)
    elif tag in ("ul", "ol"):
        for i, li in enumerate(node.find_all("li", recursive=False)):
            prefix = "• " if tag == "ul" else f"{i + 1}. "
            row = _write_paragraph(ws, row, prefix + li.get_text(strip=True))
    else:
        for child in node.children:
            row = _process_node(ws, child, row)
    return row


def _auto_col_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                for line in str(cell.value).split("\n"):
                    max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def build_excel(pages: list[tuple[int, str, str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Confluence Docs"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    row = 1
    for depth, title, html_body in pages:
        row = _write_heading(ws, row, title, depth)
        soup = BeautifulSoup(html_body, "html.parser")
        soup = clean(soup)
        for child in soup.children:
            row = _process_node(ws, child, row)
        row += 1
    return wb


def main():
    if not ROOT_PAGE_ID:
        print("[ERROR] CONFLUENCE_ROOT_PAGE_ID 환경변수를 설정하세요.")
        return
    print(f"페이지 수집 중... (root={ROOT_PAGE_ID})\n")
    pages = collect_pages(ROOT_PAGE_ID)
    print(f"\n총 {len(pages)}개 페이지 수집. Excel 생성 중...")
    wb = build_excel(pages)
    wb.save(OUTPUT_FILE)
    print(f"완료: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
