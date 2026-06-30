# converters/to_excel.py
from __future__ import annotations

from bs4 import BeautifulSoup, NavigableString, Tag
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from html_cleaner import clean, is_code_table, extract_code_lines

_H_FILL  = PatternFill("solid", fgColor="2F5496")
_H_FONT  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
_SH_FILL = PatternFill("solid", fgColor="D9E1F2")
_SH_FONT = Font(name="맑은 고딕", bold=True, color="1F3864", size=10)
_B_FONT  = Font(name="맑은 고딕", size=10)
_C_FONT  = Font(name="Consolas", size=9, color="333333")
_C_FILL  = PatternFill("solid", fgColor="F4F4F4")
_THIN    = Side(style="thin", color="AAAAAA")
_BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP    = Alignment(wrap_text=True, vertical="top")
_TOTAL_COLS = 4


def _heading(ws, row: int, text: str, depth: int) -> int:
    fill = _H_FILL if depth <= 1 else _SH_FILL
    font = _H_FONT if depth <= 1 else _SH_FONT
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill; c.font = font; c.alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_TOTAL_COLS)
    return row + 1


def _write_table(ws, row: int, tbl: Tag) -> int:
    for ri, tr in enumerate(tbl.find_all("tr")):
        col = 1
        for ct in tr.find_all(["td","th"]):
            text = ct.get_text(strip=True).replace("\u00a0"," ")
            cs = int(ct.get("colspan",1))
            is_hdr = ct.name == "th" or ri == 0
            c = ws.cell(row=row, column=col, value=text)
            c.font = _H_FONT if is_hdr else _B_FONT
            c.fill = _SH_FILL if is_hdr else PatternFill()
            c.border = _BORDER; c.alignment = _WRAP
            if cs > 1:
                ws.merge_cells(start_row=row,start_column=col,
                               end_row=row,end_column=col+cs-1)
            col += cs
        row += 1
    return row + 1


def _write_code(ws, row: int, text: str) -> int:
    for line in text.split("\n"):
        c = ws.cell(row=row, column=1, value=line)
        c.font = _C_FONT; c.fill = _C_FILL
        c.alignment = Alignment(wrap_text=False, vertical="top")
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=_TOTAL_COLS)
        row += 1
    return row + 1


def _write_para(ws, row: int, text: str) -> int:
    if not text.strip(): return row
    c = ws.cell(row=row, column=1, value=text)
    c.font = _B_FONT; c.alignment = _WRAP
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=_TOTAL_COLS)
    return row + 1


def _process(ws, node, row: int) -> int:
    if isinstance(node, NavigableString): return row
    tag = node.name
    if not tag: return row
    if tag in ("h1","h2","h3","h4","h5","h6"):
        row = _heading(ws, row, node.get_text(strip=True), int(tag[1])-1)
    elif tag == "p":
        row = _write_para(ws, row, node.get_text().replace("\u00a0"," ").strip())
    elif tag in ("pre","code"):
        row = _write_code(ws, row, node.get_text())
    elif tag == "table":
        if is_code_table(node): row = _write_code(ws, row, extract_code_lines(node))
        else: row = _write_table(ws, row, node)
    elif tag in ("ul","ol"):
        for i, li in enumerate(node.find_all("li",recursive=False)):
            prefix = "• " if tag=="ul" else f"{i+1}. "
            row = _write_para(ws, row, prefix + li.get_text(strip=True))
    else:
        for c in node.children: row = _process(ws, c, row)
    return row


def convert(pages: list[tuple[int,str,str]], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Confluence Docs"
    ws.sheet_view.showGridLines = False
    for i in range(1, _TOTAL_COLS+1):
        ws.column_dimensions[get_column_letter(i)].width = 55 if i==1 else 18

    row = 1
    for depth, title, html_body in pages:
        row = _heading(ws, row, title, depth)
        soup = clean(BeautifulSoup(html_body, "html.parser"))
        for c in soup.children: row = _process(ws, c, row)
        row += 1
    wb.save(output_path)
